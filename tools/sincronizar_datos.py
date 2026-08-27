# -*- coding: utf-8 -*-
"""
Regenera js/data.js a partir de la hoja TRAMOS de CALCULADORA_TARIFAS_2026.xlsx.

La hoja TRAMOS es la fuente unica de verdad de las tarifas (ver CLAUDE.md).
Este script evita que el Excel y la web queden descuadrados: copia los valores
con la precision completa que tiene la hoja, sin truncar decimales.

Uso:  python tools/sincronizar_datos.py
"""

import json
import os
import re

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(RAIZ, "docs", "CALCULADORA_TARIFAS_2026.xlsx")
DATA_JS = os.path.join(RAIZ, "js", "data.js")


def tipos_actuales():
    """Recupera el campo 'tipo' del data.js existente, indexado por ruta.

    La hoja TRAMOS no tiene esa columna, asi que se conserva lo que ya estaba
    en la web en vez de perderlo al regenerar. Si algun dia TRAMOS incorpora
    la columna, hay que leerla desde ahi y borrar esta funcion.
    """
    if not os.path.exists(DATA_JS):
        return {}
    texto = open(DATA_JS, encoding="utf-8").read()
    m = re.search(r"const TARIFAS_DATA = (\{.*\});", texto, re.S)
    if not m:
        return {}
    try:
        previo = json.loads(m.group(1))
    except json.JSONDecodeError:
        return {}
    return {s["ruta"]: s.get("tipo") for s in previo.get("segments", []) if s.get("tipo")}


def leer_tramos():
    tipos = tipos_actuales()
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["TRAMOS"]

    segmentos = []
    fila = 2
    while ws.cell(row=fila, column=1).value:
        ruta = ws.cell(row=fila, column=1).value
        segmentos.append({
            "ruta": ruta,
            "origen": ws.cell(row=fila, column=2).value,
            "destino": ws.cell(row=fila, column=3).value,
            "transportador": ws.cell(row=fila, column=9).value,
            "tipo": tipos.get(ruta, "Punto"),
            "fijos": ws.cell(row=fila, column=4).value,
            "variables": ws.cell(row=fila, column=5).value,
            "aom": ws.cell(row=fila, column=6).value,
        })
        fila += 1

    estampillas = []
    fila = 2
    while ws.cell(row=fila, column=12).value:
        estampillas.append({
            "transportador": ws.cell(row=fila, column=12).value,
            "fijos": ws.cell(row=fila, column=13).value,
            "variables": ws.cell(row=fila, column=14).value,
            "aom": ws.cell(row=fila, column=15).value,
        })
        fila += 1

    nodos = sorted({s["origen"] for s in segmentos} | {s["destino"] for s in segmentos})
    return {"segments": segmentos, "estampillas": estampillas, "nodes": nodos}


def main():
    datos = leer_tramos()
    cuerpo = json.dumps(datos, ensure_ascii=False, indent=2)
    contenido = (
        "// Datos de tramos y tarifas vigentes del SNT.\n"
        "// GENERADO por tools/sincronizar_datos.py desde la hoja TRAMOS de\n"
        "// docs/CALCULADORA_TARIFAS_2026.xlsx. No editar a mano: actualiza la hoja\n"
        "// TRAMOS y vuelve a ejecutar el script, para que el Excel y la web no\n"
        "// queden descuadrados.\n"
        f"const TARIFAS_DATA = {cuerpo};\n"
    )
    with open(DATA_JS, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(contenido)

    print(f"js/data.js regenerado: {len(datos['segments'])} tramos, "
          f"{len(datos['estampillas'])} estampillas, {len(datos['nodes'])} nodos")


if __name__ == "__main__":
    main()
