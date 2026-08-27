# -*- coding: utf-8 -*-
"""
Reconstruye CALCULADORA_TARIFAS_2026.xlsx a partir de la hoja TRAMOS.

Modelo tarifario (debe coincidir con js/graph.js):
  - Por tramo: Fijos*%CF + Variables*(1-%CF) + AOM (el AOM no se pondera).
  - Estampilla: una sola vez por transportador usado en la ruta.
  - Cuota de fomento e impuesto local: POR TRANSPORTADOR, sobre su propia base
    (sus tramos + su estampilla).
      * TGI      -> fomento e impuesto EDITABLES desde la hoja CALCULADORA.
      * PROMIGAS -> 3% fomento / 6% impuesto, fijos (automaticos).
  - Origen y Destino en redes distintas -> ruta COMBINADA: se suman las dos
    rutas hasta el punto de entrada de cada red, con ambas estampillas, ambos
    fomentos y ambos impuestos.

Qué toca y qué NO toca:
  - Reescribe por completo la hoja CALCULADORA.
  - Reescribe la hoja Calculo entera, incluida la MATRIZ DE PERTENENCIA, que se
    recalcula con BFS desde la hoja TRAMOS (antes estaba escrita a mano).
  - Reescribe las columnas de nodos de la hoja `listas`.
  - NO toca la hoja TRAMOS: es la fuente de verdad de las tarifas.

Uso:  python tools/actualizar_excel.py
"""

import os
import shutil
from collections import deque

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(RAIZ, "docs", "CALCULADORA_TARIFAS_2026.xlsx")

# Configuracion regulatoria por transportador. Debe coincidir con
# CONFIG_TRANSPORTADOR en js/graph.js.
#   editable=True  -> usa los porcentajes que el usuario elige en CALCULADORA
#   editable=False -> usa los porcentajes fijos de Calculo!B10 y Calculo!B11
CONFIG = {
    "TGI": {"editable": True, "entrada": "Ballenas_tgi"},
    "PROMIGAS": {"editable": False, "entrada": "Ballenas_prom"},
}
ORDEN_TRANSPORTADORES = ["TGI", "PROMIGAS"]

FOMENTO_FIJO = 0.03      # PROMIGAS
IMPUESTO_FIJO = 0.06     # PROMIGAS

# ---- Estilos ----
NAVY = "1F3864"
AZUL_CLARO = "DDE5F2"
GRIS = "F2F4F8"
NARANJA = "E8792B"

F_TITULO = Font(bold=True, size=14, color="FFFFFF")
F_SECCION = Font(bold=True, size=11, color="FFFFFF")
F_ETIQUETA = Font(bold=True, size=10)
F_CABECERA = Font(bold=True, size=9, color=NAVY)
F_NOTA = Font(italic=True, size=8, color="5B6478")
F_TOTAL = Font(bold=True, size=11, color=NAVY)

R_TITULO = PatternFill("solid", fgColor=NAVY)
R_SECCION = PatternFill("solid", fgColor="2E5395")
R_CABECERA = PatternFill("solid", fgColor=AZUL_CLARO)
R_INPUT = PatternFill("solid", fgColor="FFF6EC")
R_AUTO = PatternFill("solid", fgColor=GRIS)

BORDE = Border(*[Side(style="thin", color="C8D1E0")] * 4)

FMT_COP = '#,##0.00'
FMT_USD = '$#,##0.0000'
FMT_PCT = '0.00%'


# --------------------------------------------------------------------------
# Lectura de la fuente de datos
# --------------------------------------------------------------------------
def leer_tramos(wb):
    ws = wb["TRAMOS"]
    tramos = []
    fila = 2
    while ws.cell(row=fila, column=10).value:      # columna J = numero de tramo
        tramos.append({
            "fila": fila,
            "ruta": ws.cell(row=fila, column=1).value,
            "origen": ws.cell(row=fila, column=2).value,
            "destino": ws.cell(row=fila, column=3).value,
            "transportador": ws.cell(row=fila, column=9).value,
        })
        fila += 1
    return tramos


def construir_matriz(tramos):
    """Para cada nodo, que tramos hay entre el y el punto de entrada de su red.

    Devuelve (nodos_ordenados, red_por_nodo, flags) donde flags[nodo] es una
    lista de 0/1 del largo de `tramos`. La red se identifica por su indice en
    ORDEN_TRANSPORTADORES, igual que antes (0 = TGI, 1 = PROMIGAS).
    """
    ady = {}
    for i, t in enumerate(tramos):
        ady.setdefault(t["origen"], []).append((t["destino"], i))
        ady.setdefault(t["destino"], []).append((t["origen"], i))

    flags = {}
    red = {}
    for id_red, nombre in enumerate(ORDEN_TRANSPORTADORES):
        raiz = CONFIG[nombre]["entrada"]
        if raiz not in ady:
            raise SystemExit(f"El punto de entrada {raiz!r} de {nombre} no existe en TRAMOS")

        camino = {raiz: []}
        red[raiz] = id_red
        cola = deque([raiz])
        while cola:
            actual = cola.popleft()
            for vecino, i_tramo in ady[actual]:
                if vecino not in camino:
                    camino[vecino] = camino[actual] + [i_tramo]
                    red[vecino] = id_red
                    cola.append(vecino)

        for nodo, tramos_hasta_raiz in camino.items():
            fila = [0] * len(tramos)
            for i in tramos_hasta_raiz:
                fila[i] = 1
            flags[nodo] = fila

    sueltos = {t["origen"] for t in tramos} | {t["destino"] for t in tramos}
    sueltos -= set(flags)
    if sueltos:
        raise SystemExit(
            "Estos nodos no cuelgan de ningun punto de entrada conocido: "
            + ", ".join(sorted(sueltos))
            + ". Revisa CONFIG en este script."
        )

    return sorted(flags), red, flags


# --------------------------------------------------------------------------
# Utilidades de escritura
# --------------------------------------------------------------------------
def limpiar_hoja(ws, col_fin):
    # Deshacer las fusiones primero: en openpyxl las celdas de un rango
    # fusionado son de solo lectura (ver docs/PROBLEMAS_Y_SOLUCIONES.md P1).
    for rango in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rango))
    for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=col_fin):
        for celda in fila:
            celda.value = None
            celda.fill = PatternFill()
            celda.font = Font()
            celda.border = Border()
            celda.number_format = "General"


def etiqueta(ws, coord, texto, font=F_ETIQUETA):
    c = ws[coord]
    c.value = texto
    c.font = font
    return c


# --------------------------------------------------------------------------
# Hoja Calculo (motor)
# --------------------------------------------------------------------------
def construir_calculo(wb, tramos, nodos, red, flags):
    ws = wb["Calculo"]
    limpiar_hoja(ws, 40)

    n = len(tramos)
    n_transp = len(ORDEN_TRANSPORTADORES)

    # Distribucion de filas, derivada del numero de tramos y transportadores:
    # si manana hay mas tramos, todo se corre solo.
    f_tramo_ini = 21
    f_tramo_fin = f_tramo_ini + n - 1
    f_tramo_tot = f_tramo_fin + 1
    f_desg_tit = f_tramo_tot + 2
    f_desg_cab = f_desg_tit + 1
    f_desg_ini = f_desg_cab + 1
    f_desg_tot = f_desg_ini + n_transp
    f_res_tit = f_desg_tot + 2
    f_res_ini = f_res_tit + 1
    f_texto = f_res_ini + 8 + 1
    f_mat_tit = f_texto + 3
    f_mat_cab = f_mat_tit + 1
    f_mat_ini = f_mat_cab + 1
    f_mat_fin = f_mat_ini + len(nodos) - 1

    col_mat_ini = 3                              # columna C
    col_mat_fin = col_mat_ini + n - 1
    MAT = (f"$C${f_mat_ini}:${get_column_letter(col_mat_fin)}${f_mat_fin}")
    NODOS = f"$A${f_mat_ini}:$A${f_mat_fin}"
    REDES = f"$B${f_mat_ini}:$B${f_mat_fin}"

    ws["A1"] = "MOTOR DE CALCULO - no editar (alimenta la hoja CALCULADORA)"
    ws["A1"].font = Font(bold=True, color=NAVY)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22

    # --- Parametros ---
    etiqueta(ws, "A3", "PARAMETROS DE ENTRADA", F_TOTAL)
    params = [
        (4, "Origen", "=CALCULADORA!T5", None),
        (5, "Destino", "=CALCULADORA!T6", None),
        (6, "%CF (fraccion Fijo)", "=CALCULADORA!T7", FMT_PCT),
        (7, "TRM (COP/USD)", "=CALCULADORA!T8", FMT_COP),
        (8, "% Impuesto local TGI (editable)", "=CALCULADORA!T9", FMT_PCT),
        (9, "% Cuota de fomento TGI (editable)", "=CALCULADORA!T10", FMT_PCT),
        (10, "% Cuota de fomento PROMIGAS (fijo)", FOMENTO_FIJO, FMT_PCT),
        (11, "% Impuesto local PROMIGAS (fijo)", IMPUESTO_FIJO, FMT_PCT),
    ]
    for fila, label, valor, fmt in params:
        etiqueta(ws, f"A{fila}", label)
        c = ws[f"B{fila}"]
        c.value = valor
        if fmt:
            c.number_format = fmt
        if fila in (10, 11):
            c.fill = R_INPUT

    # --- Validacion / tipo de ruta ---
    etiqueta(ws, "A13", "VALIDACION DE RUTA", F_TOTAL)
    etiqueta(ws, "A14", "Fila de Origen en matriz")
    ws["B14"] = f"=MATCH(B4,{NODOS},0)"
    etiqueta(ws, "A15", "Fila de Destino en matriz")
    ws["B15"] = f"=MATCH(B5,{NODOS},0)"
    etiqueta(ws, "A16", "Red de Origen")
    ws["B16"] = f"=INDEX({REDES},B14)"
    etiqueta(ws, "A17", "Red de Destino")
    ws["B17"] = f"=INDEX({REDES},B15)"
    etiqueta(ws, "A18", "Ruta valida?")
    ws["B18"] = '=IF(B4=B5,"ORIGEN = DESTINO","SI")'
    etiqueta(ws, "A19", "Tipo de ruta")
    ws["B19"] = ('=IF($B$18<>"SI","-",IF($B$16=$B$17,"DIRECTA",'
                 '"COMBINADA - se suman las dos rutas hasta el punto de entrada de cada red"))')

    # --- Tabla de tramos ---
    cabeceras = ["Tramo", "Ruta", "Transportador", "Fijos", "Variables", "AOM",
                 "Incluido", "Fijo ponderado", "Variable ponderado", "AOM incluido"]
    for i, texto in enumerate(cabeceras, start=1):
        c = ws.cell(row=f_tramo_ini - 1, column=i, value=texto)
        c.font, c.fill = F_CABECERA, R_CABECERA

    for k, tramo in enumerate(tramos):
        fila = f_tramo_ini + k
        ft = tramo["fila"]
        col = col_mat_ini + k
        ws[f"A{fila}"] = f"=TRAMOS!J{ft}"
        ws[f"B{fila}"] = f"=TRAMOS!A{ft}"
        ws[f"C{fila}"] = f"=TRAMOS!I{ft}"
        ws[f"D{fila}"] = f"=TRAMOS!D{ft}"
        ws[f"E{fila}"] = f"=TRAMOS!E{ft}"
        ws[f"F{fila}"] = f"=TRAMOS!F{ft}"
        # La matriz guarda, por nodo, los tramos hasta la raiz de SU red.
        #  - Misma red      -> el camino Origen-Destino es ABS(fo - fd) (XOR).
        #  - Redes distintas -> ruta combinada: la union es fo + fd, porque las
        #    redes son disjuntas y cada camino va hacia una raiz distinta.
        fo = f"INDEX({MAT},$B$14,{k + 1})"
        fd = f"INDEX({MAT},$B$15,{k + 1})"
        ws[f"G{fila}"] = (f'=IF($B$18<>"SI",0,IF($B$16=$B$17,'
                          f'ABS({fo}-{fd}),{fo}+{fd}))')
        ws[f"H{fila}"] = f"=D{fila}*$B$6*G{fila}"
        ws[f"I{fila}"] = f"=E{fila}*(1-$B$6)*G{fila}"
        ws[f"J{fila}"] = f"=F{fila}*G{fila}"
        for c in "DEFHIJ":
            ws[f"{c}{fila}"].number_format = FMT_COP

    etiqueta(ws, f"B{f_tramo_tot}", "TOTALES", F_TOTAL)
    for c in "HIJ":
        celda = ws[f"{c}{f_tramo_tot}"]
        celda.value = f"=SUM({c}{f_tramo_ini}:{c}{f_tramo_fin})"
        celda.number_format, celda.font = FMT_COP, F_TOTAL

    # --- Desglose por transportador ---
    etiqueta(ws, f"A{f_desg_tit}", "DESGLOSE POR TRANSPORTADOR", F_TOTAL)
    cab = ["Transportador", "Usa", "Fijos pond.", "Variables pond.", "AOM",
           "Estampilla", "Base", "% Fomento", "Fomento", "% Impuesto",
           "Impuesto", "Total"]
    for i, texto in enumerate(cab, start=1):
        c = ws.cell(row=f_desg_cab, column=i, value=texto)
        c.font, c.fill = F_CABECERA, R_CABECERA

    r_transp = f"$C${f_tramo_ini}:$C${f_tramo_fin}"
    r_incl = f"$G${f_tramo_ini}:$G${f_tramo_fin}"
    r_fijo = f"$H${f_tramo_ini}:$H${f_tramo_fin}"
    r_var = f"$I${f_tramo_ini}:$I${f_tramo_fin}"
    r_aom = f"$J${f_tramo_ini}:$J${f_tramo_fin}"

    for i, nombre in enumerate(ORDEN_TRANSPORTADORES):
        f = f_desg_ini + i
        editable = CONFIG[nombre]["editable"]
        celda_fom = "$B$9" if editable else "$B$10"
        celda_imp = "$B$8" if editable else "$B$11"

        ws[f"A{f}"] = nombre
        ws[f"A{f}"].font = F_ETIQUETA
        ws[f"B{f}"] = f'=IF(SUMPRODUCT(({r_transp}=$A{f})*{r_incl})>0,"SI","NO")'
        ws[f"C{f}"] = f"=SUMPRODUCT(({r_transp}=$A{f})*{r_fijo})"
        ws[f"D{f}"] = f"=SUMPRODUCT(({r_transp}=$A{f})*{r_var})"
        ws[f"E{f}"] = f"=SUMPRODUCT(({r_transp}=$A{f})*{r_aom})"
        # SUMIF devuelve 0 si el transportador no tiene fila de estampilla.
        ws[f"F{f}"] = (
            f'=IF($B{f}="SI",'
            f'SUMIF(TRAMOS!$L$2:$L$3,$A{f},TRAMOS!$M$2:$M$3)*$B$6'
            f'+SUMIF(TRAMOS!$L$2:$L$3,$A{f},TRAMOS!$N$2:$N$3)*(1-$B$6)'
            f'+SUMIF(TRAMOS!$L$2:$L$3,$A{f},TRAMOS!$O$2:$O$3),0)'
        )
        ws[f"G{f}"] = f"=C{f}+D{f}+E{f}+F{f}"
        ws[f"H{f}"] = f"={celda_fom}"
        ws[f"I{f}"] = f"=G{f}*H{f}"
        ws[f"J{f}"] = f"={celda_imp}"
        ws[f"K{f}"] = f"=G{f}*J{f}"
        ws[f"L{f}"] = f"=G{f}+I{f}+K{f}"
        for c in "CDEFGIKL":
            ws[f"{c}{f}"].number_format = FMT_COP
        for c in "HJ":
            ws[f"{c}{f}"].number_format = FMT_PCT

    etiqueta(ws, f"A{f_desg_tot}", "TOTALES", F_TOTAL)
    for c in "CDEFGIKL":
        celda = ws[f"{c}{f_desg_tot}"]
        celda.value = f"=SUM({c}{f_desg_ini}:{c}{f_desg_tot - 1})"
        celda.number_format, celda.font = FMT_COP, F_TOTAL

    # --- Resultados ---
    etiqueta(ws, f"A{f_res_tit}", "RESULTADOS", F_TOTAL)
    b_base = f"B{f_res_ini + 2}"
    b_total = f"B{f_res_ini + 5}"
    resultados = [
        ("Subtotal tramos (COP)",
         f"=H{f_tramo_tot}+I{f_tramo_tot}+J{f_tramo_tot}", FMT_COP),
        ("Total estampillas (COP)", f"=F{f_desg_tot}", FMT_COP),
        ("Costo base transporte (COP/KPC)", f"=G{f_desg_tot}", FMT_COP),
        ("Cuota de fomento total (COP/KPC)", f"=I{f_desg_tot}", FMT_COP),
        ("Impuesto local total (COP/KPC)", f"=K{f_desg_tot}", FMT_COP),
        ("TOTAL (COP/KPC)", f"=L{f_desg_tot}", FMT_COP),
        ("Costo base transporte (USD/KPC)", f"=IF(B7=0,0,{b_base}/B7)", FMT_USD),
        ("TOTAL (USD/KPC)", f"=IF(B7=0,0,{b_total}/B7)", FMT_USD),
    ]
    for i, (label, formula, fmt) in enumerate(resultados):
        fila = f_res_ini + i
        etiqueta(ws, f"A{fila}", label)
        c = ws[f"B{fila}"]
        c.value = formula
        c.number_format = fmt
        if "TOTAL (" in label:
            c.font = F_TOTAL

    # --- Texto de la ruta ---
    # TEXTJOIN necesita el prefijo _xlfn. y evaluacion de arreglo cuando el
    # archivo lo genera openpyxl (ver docs/PROBLEMAS_Y_SOLUCIONES.md P2).
    etiqueta(ws, f"A{f_texto}", "Tramos utilizados (texto)")
    ws[f"B{f_texto}"] = ArrayFormula(
        f"B{f_texto}",
        f'=IF(B18<>"SI","",_xlfn.TEXTJOIN(" -> ",TRUE,'
        f'IF($G${f_tramo_ini}:$G${f_tramo_fin}=1,'
        f'$B${f_tramo_ini}:$B${f_tramo_fin},"")))'
    )

    # --- Matriz de pertenencia (recalculada con BFS, no escrita a mano) ---
    ws[f"A{f_mat_tit}"] = ("MATRIZ DE PERTENENCIA (tramos entre el nodo y el punto de "
                           "entrada de su red) - generada por tools/actualizar_excel.py")
    ws[f"A{f_mat_tit}"].font = Font(bold=True, color=NAVY)
    ws.cell(row=f_mat_cab, column=1, value="Nodo").font = F_CABECERA
    ws.cell(row=f_mat_cab, column=1).fill = R_CABECERA
    ws.cell(row=f_mat_cab, column=2, value="Red").font = F_CABECERA
    ws.cell(row=f_mat_cab, column=2).fill = R_CABECERA
    for k in range(n):
        c = ws.cell(row=f_mat_cab, column=col_mat_ini + k, value=k + 1)
        c.font, c.fill = F_CABECERA, R_CABECERA

    for i, nodo in enumerate(nodos):
        fila = f_mat_ini + i
        ws.cell(row=fila, column=1, value=nodo)
        ws.cell(row=fila, column=2, value=red[nodo])
        for k, valor in enumerate(flags[nodo]):
            ws.cell(row=fila, column=col_mat_ini + k, value=valor)

    return {
        "f_ruta_valida": 18, "f_tipo_ruta": 19,
        "f_desg_ini": f_desg_ini, "f_desg_tot": f_desg_tot,
        "f_res_ini": f_res_ini, "f_texto": f_texto,
    }


# --------------------------------------------------------------------------
# Hoja listas (fuente de los desplegables)
# --------------------------------------------------------------------------
def construir_listas(wb, nodos):
    ws = wb["listas"]
    # solo se reescriben las dos columnas de nodos; el resto de la hoja
    # (combos de %CF e impuesto) se deja como esta.
    for fila in range(2, ws.max_row + 2):
        ws.cell(row=fila, column=1).value = None
        ws.cell(row=fila, column=2).value = None
    ws["A1"], ws["B1"] = "Origenes", "Destinos"
    for i, nodo in enumerate(nodos):
        ws.cell(row=2 + i, column=1, value=nodo)
        ws.cell(row=2 + i, column=2, value=nodo)
    return len(nodos) + 1     # ultima fila con datos


# --------------------------------------------------------------------------
# Hoja CALCULADORA (interfaz)
# --------------------------------------------------------------------------
def construir_calculadora(wb, ref, fila_final_listas):
    idx = wb.sheetnames.index("CALCULADORA")
    del wb["CALCULADORA"]
    ws = wb.create_sheet("CALCULADORA", idx)

    anchos = {"A": 26, "B": 12, "C": 16, "D": 14, "E": 14, "F": 14,
              "G": 12, "H": 12, "I": 14, "J": 12, "K": 14, "L": 16}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["T"].hidden = True   # columna puente

    ws.merge_cells("A1:L2")
    t = ws["A1"]
    t.value = "CALCULADORA DE TARIFAS DE TRANSPORTE DE GAS - SNT"
    t.font, t.fill = F_TITULO, R_TITULO
    t.alignment = Alignment(horizontal="center", vertical="center")

    def seccion(fila, texto):
        ws.merge_cells(f"A{fila}:L{fila}")
        c = ws[f"A{fila}"]
        c.value = texto
        c.font, c.fill = F_SECCION, R_SECCION
        c.alignment = Alignment(vertical="center")

    seccion(4, "1. SELECCIONE LA RUTA Y LAS OPCIONES")
    entradas = [
        (5, "Origen", "Jobo", None, "=C5"),
        (6, "Destino", "Ballenas_prom", None, "=C6"),
        # mismo arranque que la web: 0/100 (ver docs/DECISIONES.md D24)
        (7, "Opcion %CF (Fijo/Variable)", "0% Fijo / 100% Variable",
         None, "=VLOOKUP(C7,listas!$D$2:$E$8,2,0)"),
        (8, "TRM (COP por USD)", 4000, FMT_COP, "=C8"),
        # mismo valor de arranque que la web: 6% (ver docs/DECISIONES.md D18)
        (9, "Impuesto local TGI", "6% - Impuesto especial municipal",
         None, "=VLOOKUP(C9,listas!$G$2:$H$4,2,0)"),
        (10, "Cuota de fomento TGI", 0.03, FMT_PCT, "=C10"),
    ]
    for fila, label, valor, fmt, puente in entradas:
        ws.merge_cells(f"A{fila}:B{fila}")
        etiqueta(ws, f"A{fila}", label)
        ws.merge_cells(f"C{fila}:D{fila}")
        c = ws[f"C{fila}"]
        c.value = valor
        c.fill, c.border = R_INPUT, BORDE
        if fmt:
            c.number_format = fmt
        ws[f"T{fila}"] = puente

    ws.merge_cells("E9:L10")
    n = ws["E9"]
    n.value = ("Impuesto local y cuota de fomento de arriba aplican SOLO A TGI. "
               "PROMIGAS se liquida automaticamente con 3% de cuota de fomento "
               "y 6% de impuesto local (no editables).")
    n.font = Font(bold=True, size=9, color=NARANJA)
    n.alignment = Alignment(wrap_text=True, vertical="center")
    n.fill = R_INPUT

    seccion(12, "2. RUTA CALCULADA")
    ws.merge_cells("A13:B13"); etiqueta(ws, "A13", "Tipo de ruta:")
    ws.merge_cells("C13:L13")
    ws["C13"] = (f'=Calculo!B{ref["f_ruta_valida"]}&'
                 f'IF(Calculo!B{ref["f_tipo_ruta"]}="-",""," / "&Calculo!B{ref["f_tipo_ruta"]})')
    ws.merge_cells("A14:B14"); etiqueta(ws, "A14", "Tramos utilizados:")
    ws.merge_cells("C14:L14"); ws["C14"] = f'=Calculo!B{ref["f_texto"]}'
    ws["C14"].alignment = Alignment(wrap_text=True)

    seccion(16, "3. DESGLOSE POR TRANSPORTADOR (cada uno liquida su estampilla, su fomento y su impuesto)")
    cabeceras = ["Transportador", "Usa", "Fijos pond.", "Variables pond.", "AOM",
                 "Estampilla", "Base", "% Fomento", "Fomento", "% Impuesto",
                 "Impuesto", "Total"]
    for i, texto in enumerate(cabeceras, start=1):
        c = ws.cell(row=17, column=i, value=texto)
        c.font, c.fill, c.border = F_CABECERA, R_CABECERA, BORDE
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, nombre in enumerate(ORDEN_TRANSPORTADORES):
        fila_ws = 18 + i
        fila_calc = ref["f_desg_ini"] + i
        modo = "editable" if CONFIG[nombre]["editable"] else "automatico"
        c = ws.cell(row=fila_ws, column=1, value=f"{nombre} ({modo})")
        c.font, c.border = F_ETIQUETA, BORDE
        if not CONFIG[nombre]["editable"]:
            c.fill = R_AUTO
        for col in range(2, 13):
            letra = get_column_letter(col)
            cel = ws.cell(row=fila_ws, column=col, value=f"=Calculo!{letra}{fila_calc}")
            cel.border = BORDE
            if col in (8, 10):
                cel.number_format = FMT_PCT
            elif col >= 3:
                cel.number_format = FMT_COP

    fila_tot_ws = 18 + len(ORDEN_TRANSPORTADORES)
    c = ws.cell(row=fila_tot_ws, column=1, value="TOTALES")
    c.font, c.border = F_TOTAL, BORDE
    for col in range(2, 13):
        letra = get_column_letter(col)
        cel = ws.cell(row=fila_tot_ws, column=col)
        cel.border = BORDE
        if col != 2 and col not in (8, 10):
            cel.value = f'=Calculo!{letra}{ref["f_desg_tot"]}'
            cel.number_format, cel.font = FMT_COP, F_TOTAL

    f0 = fila_tot_ws + 2
    seccion(f0, "4. RESULTADO FINAL")
    totales = [
        ("Costo base transporte (COP/KPC)", 2, FMT_COP),
        ("Cuota de fomento - suma transportadores (COP/KPC)", 3, FMT_COP),
        ("Impuesto local - suma transportadores (COP/KPC)", 4, FMT_COP),
        ("TOTAL (COP/KPC)", 5, FMT_COP),
        ("Costo base transporte (USD/KPC)", 6, FMT_USD),
        ("TOTAL (USD/KPC)", 7, FMT_USD),
    ]
    for i, (label, desplazamiento, fmt) in enumerate(totales):
        fila = f0 + 1 + i
        ws.merge_cells(f"A{fila}:D{fila}")
        etiqueta(ws, f"A{fila}", label)
        ws.merge_cells(f"E{fila}:F{fila}")
        c = ws[f"E{fila}"]
        c.value = f'=Calculo!B{ref["f_res_ini"] + desplazamiento}'
        c.number_format = fmt
        c.border = BORDE
        if label.startswith("TOTAL"):
            c.font = F_TOTAL
            ws[f"A{fila}"].font = F_TOTAL

    f_nota = f0 + len(totales) + 3
    ws.merge_cells(f"A{f_nota}:L{f_nota + 3}")
    n = ws[f"A{f_nota}"]
    n.value = (
        "Notas: %CF = porcentaje que se cobra como cargo fijo (el resto se cobra como cargo "
        "variable); el AOM se cobra 100% independiente del %CF. La estampilla se cobra una sola "
        "vez por transportador usado en la ruta. Si Origen y Destino estan en redes distintas, la "
        "ruta es COMBINADA: se suma la ruta del Origen hasta el punto de entrada de su red y la "
        "del punto de entrada de la red del Destino hasta el Destino, y se cobran las estampillas, "
        "cuotas de fomento e impuestos de ambos transportadores. La cuota de fomento y el impuesto "
        "local se liquidan POR TRANSPORTADOR sobre su propia base (sus tramos + su estampilla): "
        "TGI usa los porcentajes editables de arriba; PROMIGAS usa 3% y 6% fijos, que se cambian "
        "en Calculo!B10 y Calculo!B11. Fuente de tarifas: hoja TRAMOS."
    )
    n.font = F_NOTA
    n.alignment = Alignment(wrap_text=True, vertical="top")

    for celda, origen in [("C5", f"listas!$A$2:$A${fila_final_listas}"),
                          ("C6", f"listas!$B$2:$B${fila_final_listas}"),
                          ("C7", "listas!$D$2:$D$8"),
                          ("C9", "listas!$G$2:$G$4")]:
        dv = DataValidation(type="list", formula1=origen, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[celda])

    ws.sheet_view.showGridLines = False
    return ws


def main():
    if not os.path.exists(XLSX):
        raise SystemExit(f"No se encontro {XLSX}")

    wb = openpyxl.load_workbook(XLSX)
    tramos = leer_tramos(wb)
    nodos, red, flags = construir_matriz(tramos)

    ref = construir_calculo(wb, tramos, nodos, red, flags)
    fila_final_listas = construir_listas(wb, nodos)
    construir_calculadora(wb, ref, fila_final_listas)

    wb.active = wb.sheetnames.index("CALCULADORA")
    wb.save(XLSX)

    resumen = ", ".join(
        f"{t}: {sum(1 for x in tramos if x['transportador'] == t)} tramos"
        for t in ORDEN_TRANSPORTADORES
    )
    print(f"Actualizado {os.path.basename(XLSX)} - {len(tramos)} tramos "
          f"({resumen}), {len(nodos)} nodos")


if __name__ == "__main__":
    main()
